import os
from pathlib import Path
import timeit
from openpyxl import load_workbook

challenge_dir = Path(os.path.dirname(__file__)).parent
solution_file = challenge_dir / "solutions/solution.py"
excel_file = challenge_dir / "solutions/Standings.xlsx"
assert solution_file.exists(), "No solution file found"

if __name__ == "__main__":
    from challenge5_performances.solutions.solution import run, leagues

    start = timeit.default_timer()
    run()
    exec_time = timeit.default_timer() - start

    assert exec_time < 1.5, f"Execution time = {exec_time}"

    assert excel_file.exists(), "No excel file found"

    wb = load_workbook(filename=excel_file)
    assert wb.sheetnames == [league.name for league in leagues], "Incorrect sheet names"
    assert all(
        sheet.max_row > 10 for sheet in wb.worksheets
    ), "Missing data in excel file"
